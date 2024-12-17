import openpyxl
from openpyxl.styles import Alignment, NamedStyle
import pyodbc


def cargaDatos():
    # UBICANDO Y CARGANDO ARCHIVO DE EXCEL
    rutaArchivo = r"C:\ARCHIVOEXCEL.xlsx"
    libroExcel = openpyxl.load_workbook(rutaArchivo)
    hojaExcel = libroExcel['HOJADEARCHIVO']

    # DEFINIENDO ESTILOS
    centrar = Alignment(horizontal='center', vertical='center')
    date_style = NamedStyle(name="date_style", number_format='DD-MM-YYYY')
    currency_style = NamedStyle(name="currency_style",
                                number_format='"$"#,##0.00_-')

    # Agregar los estilos al libro si no existen
    if "date_style" not in libroExcel.named_styles:
        libroExcel.add_named_style(date_style)
    if "currency_style" not in libroExcel.named_styles:
        libroExcel.add_named_style(currency_style)

    ultimaFila = hojaExcel.max_row
    ultimaFactura = hojaExcel.cell(row=ultimaFila, column=1).value

    # Cargar todos los ID_DOCUMENTO existentes en el Excel
    id_documentos_excel = set()
    for fila in hojaExcel.iter_rows(min_row=2, max_row=ultimaFila, min_col=1, max_col=1, values_only=True):
        if fila[0] is not None:
            id_documentos_excel.add(fila[0])

    # Detalles de la conexión
    # Puede ser una dirección IP o un nombre de servidor
    server = 'SERVIDOR'
    database = 'BASEDEDATOS'
    username = 'USUARIO'
    password = 'PASSWORD'

    # Establecer la conexión
    conexion = pyodbc.connect(
        f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}')
    print("Conexión exitosa")

    # Crear un cursor
    cursor = conexion.cursor()

    # Ejecutar una consulta
    query = f"""
    QUERY DE SQL SERVER
    """

    try:
        cursor.execute(query)
        rows = cursor.fetchall()
        nuevasFacturas = [list(row) for row in rows]

        print("Resultados de la consulta:")
        for fila in nuevasFacturas:
            if fila[0] not in id_documentos_excel:  # Validar si el ID_DOCUMENTO no está repetido
                print(fila)
                ultimaFila += 1
                for col, valor in enumerate(fila, start=1):
                    if col == 1:
                        valor = int(valor)
                    celda = hojaExcel.cell(
                        row=ultimaFila, column=col, value=valor)
                    celda.alignment = centrar

                    # Aplicar formato de fecha a la segunda columna
                    if col == 2:
                        celda.number_format = 'DD-MM-YYYY'
                    # Aplicar formato de moneda a la cuarta columna
                    elif col == 4:
                        celda.number_format = '"$"#,##0.00_-'

        libroExcel.save(rutaArchivo)
        print("Los datos se han insertado y formateado correctamente.")

    except pyodbc.Error as err:
        print(f"Error en la ejecución de la consulta: {err}")

    finally:
        # Cerrar la conexión
        if 'conexion' in locals():
            conexion.close()
            print("Conexión cerrada")
